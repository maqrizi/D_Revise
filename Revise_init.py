# Prepare Libraries
#------------------
import openpyxl



#Open Revision Sheets 
#---------------------------------
def OpenRevSheet(DB_type):

    if (DB_type=='alt'):
        revision_book=openpyxl.load_workbook('ReviseWords.xlsx')
        master_book=openpyxl.load_workbook('OldWords.xlsx')
    elif (DB_type=='neu'):
        revision_book=openpyxl.load_workbook('ReviseWords.xlsx')
        master_book=openpyxl.load_workbook('NewWords.xlsx')
    else:
        print('Du hast falsch DB Beschreibung eingeschrieben')

    master_sheet=master_book.get_sheet_by_name('first sheet')
    revision_sheet=revision_book.create_sheet()

    k=1
    m=1
    Table_Lenght=0

    while master_sheet.cell(row=k, column=1).value!=None:
        k=k+1
        Table_Lenght=k

    for m in range(1, Table_Lenght):
        revision_sheet.cell(row=m, column=1).value=master_sheet.cell(row=m, column=1).value
        revision_sheet.cell(row=m, column=2).value=master_sheet.cell(row=m, column=2).value
        
    revision_sheet.cell(row=1, column=3).value='User Answer'
    
    return revision_sheet, revision_book, Table_Lenght

   

