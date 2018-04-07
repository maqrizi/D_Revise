# Prepare Libraries
#------------------
import openpyxl


def GetRevResult(revision_sheet, start_index, Table_Lenght):
    #Validate User Answers
    #---------------------
    rechtig = 0
    falsch = 0
    revision_sheet.cell(row=1, column=4).value='Result'
    for p in range(start_index,start_index+Table_Lenght):
        if (revision_sheet.cell(row=p, column=1).value==revision_sheet.cell(row=p, column=3).value):
            revision_sheet.cell(row=p, column=4).value='Rechtig'
            rechtig=rechtig+1
        else:
            revision_sheet.cell(row=p, column=4).value='Falsch'
            falsch=falsch+1


    #Calculate User Result
    #---------------------
    result=(rechtig/(rechtig+falsch))*100
    return result

    

