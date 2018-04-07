# Prepare Libraries
#------------------
import openpyxl
import Revise_init
import Revise_parse
import Revise_ask
import Revise_result
#from Revise_init import *


#Initialization of the Application
#---------------------------------
print('Welche Woerter willst du widerholen?alt/neu')
DB_Category=input()
if ((DB_Category == 'alt') or (DB_Category == 'neu')):
    revision_sheet, revision_book, Table_Length = Revise_init.OpenRevSheet(DB_Category)
else:
     print('Bitte Geben richtige Eingabe ein (alt oder neu)')   

#Parse Revision Sheet
#--------------------
start_index, RevWordsNo = Revise_parse.ParseRevSheet(revision_sheet, Table_Length)

#Interactive Questions
#---------------------
WordsCounter = RevWordsNo
for n in range (start_index,start_index+RevWordsNo):
    Revise_ask.RevWordsCounter = WordsCounter
    NextMsg, CellInTurn = Revise_ask.AskRevWords(revision_sheet, start_index, RevWordsNo)
    print(NextMsg)
    user_answer=input()
    revision_sheet.cell(row=CellInTurn, column=3).value=user_answer
    WordsCounter = WordsCounter -1

#Calculate Result
#----------------
result = Revise_result.GetRevResult(revision_sheet, start_index, RevWordsNo)

revision_sheet.cell(row=5, column=6).value = 'Result for today is'
revision_sheet.cell(row=6, column=6).value = result
print('result is {} %'.format(result))

   

#Save Procedure
#--------------
revision_book.save('ReviseWords.xlsx')
